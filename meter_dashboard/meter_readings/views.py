import psycopg2
import os
import sys
import json
from datetime import datetime, timedelta
from io import BytesIO
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponse
from django.shortcuts import render
from django.conf import settings
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def api_meter_readings(request):
    db_settings = getattr(settings, 'DATABASES', {}).get('default', {})
    conn = None
    cur = None
    rows = []
    columns = []
    try:
        conn = psycopg2.connect(
            dbname=db_settings.get('NAME', 'mfmdb'),
            user=db_settings.get('USER', 'mfmuser'),
            password=db_settings.get('PASSWORD', 'devi'),
            host=db_settings.get('HOST', 'localhost'),
            port=db_settings.get('PORT', 5432)
        )
        cur = conn.cursor()
        try:
            cur.execute('SELECT * FROM meter_readings ORDER BY time DESC LIMIT 10;')
            rows = cur.fetchall()
            columns = [desc[0] for desc in cur.description]
        except Exception:
            # Table may not exist yet; return empty result
            rows, columns = [], []
    finally:
        try:
            if cur is not None:
                cur.close()
        finally:
            if conn is not None:
                conn.close()
    # Return as JSON: list of dicts
    data = [dict(zip(columns, row)) for row in rows]
    return JsonResponse({'readings': data})


def _fetch_recent_alert_events(device_id=None, limit=300):
    """Import-safe helper to fetch alert events from Redis via Celery task module."""
    try:
        base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
        if base_dir not in sys.path:
            sys.path.append(base_dir)
        from iot_scripts.alerting.celery_alert_tasks import fetch_recent_alert_events
        return fetch_recent_alert_events(device_id=device_id, limit=limit)
    except Exception:
        return []


def latest_readings(request):
    # Enhanced view with filtering and time-series support
    import psycopg2
    db_settings = getattr(settings, 'DATABASES', {}).get('default', {})
    
    # Get filter parameters
    view_mode = request.GET.get('view_mode', 'latest')  # 'latest' or 'timeseries'
    selected_meters = request.GET.getlist('meters')
    hours_back = request.GET.get('hours', '24')
    
    try:
        hours_back = int(hours_back)
    except ValueError:
        hours_back = 24
    
    conn = None
    cur = None
    readings = []
    columns = []
    available_meters = []
    
    try:
        conn = psycopg2.connect(
            dbname=db_settings.get('NAME', 'mfmdb'),
            user=db_settings.get('USER', 'mfmuser'),
            password=db_settings.get('PASSWORD', 'devi'),
            host=db_settings.get('HOST', 'localhost'),
            port=db_settings.get('PORT', 5432)
        )
        cur = conn.cursor()
        
        # Get list of available meters
        try:
            cur.execute("SELECT DISTINCT meter_name FROM meter_readings ORDER BY meter_name;")
            available_meters = [row[0] for row in cur.fetchall()]
        except Exception:
            available_meters = []
        
        # Build query based on view mode
        if view_mode == 'latest':
            query = "SELECT DISTINCT ON (meter_name) * FROM meter_readings"
            if selected_meters:
                placeholders = ','.join(['%s'] * len(selected_meters))
                query += f" WHERE meter_name IN ({placeholders})"
            query += " ORDER BY meter_name, time DESC;"
            
            try:
                if selected_meters:
                    cur.execute(query, selected_meters)
                else:
                    cur.execute(query)
                readings = cur.fetchall()
                columns = [desc[0] for desc in cur.description]
            except Exception:
                readings, columns = [], []
        else:  # timeseries
            query = "SELECT * FROM meter_readings WHERE time >= NOW() - INTERVAL '%s hours'"
            params = [hours_back]
            
            if selected_meters:
                placeholders = ','.join(['%s'] * len(selected_meters))
                query += f" AND meter_name IN ({placeholders})"
                params.extend(selected_meters)
            
            query += " ORDER BY time DESC LIMIT 1000;"
            
            try:
                cur.execute(query, params)
                readings = cur.fetchall()
                columns = [desc[0] for desc in cur.description]
            except Exception:
                readings, columns = [], []
    finally:
        try:
            if cur is not None:
                cur.close()
        finally:
            if conn is not None:
                conn.close()
    
    all_rows = readings
    # Debug after computing all_rows
    try:
        print("DEBUG all_rows:", all_rows)
    except Exception:
        pass

    # Compute active alerts from Redis events (no current_alerts.json)
    events = _fetch_recent_alert_events(limit=500)
    # Determine active devices: last event per (device,param); if not 'recovered' => active
    active_devices = set()
    last_by_key = {}
    for e in events:
        key = (str(e.get('device_id')), e.get('param'))
        last_by_key[key] = e
    for (did, _), e in last_by_key.items():
        if e.get('alert_type') != 'recovered':
            active_devices.add(did)

    # Prepare rows with alert flags for easier templating
    try:
        device_id_index = columns.index('device_id')
    except ValueError:
        device_id_index = None
    rows_aug = []
    for r in all_rows:
        dev_id_val = None
        if device_id_index is not None and len(r) > device_id_index:
            try:
                dev_id_val = str(r[device_id_index])
            except Exception:
                dev_id_val = None
        rows_aug.append({
            'cells': r,
            'has_alert': (dev_id_val in active_devices)
        })

    return render(request, 'meter_readings/latest_readings.html', {
        'columns': columns,
        'rows': all_rows,
        'rows_aug': rows_aug,
        'page_title': 'Meter Readings Dashboard',
        'alert_state': {'active_alerts': {}, 'source': 'redis_events'},
        'available_meters': available_meters,
        'selected_meters': selected_meters,
        'view_mode': view_mode,
        'hours_back': hours_back,
        'excel_available': EXCEL_AVAILABLE,
    })


# Path to the shared failure modes JSON file (project-relative, inside iot_scripts)
FAILURE_MODES_FILE = os.path.abspath(os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
    'iot_scripts', 'failure_modes.json'
))


def load_failure_modes():
    try:
        with open(FAILURE_MODES_FILE, 'r') as f:
            return json.load(f)
    except Exception:
        return {}


def save_failure_modes(modes):
    try:
        with open(FAILURE_MODES_FILE, 'w') as f:
            json.dump(modes, f, indent=2)
    except Exception as e:
        print(f"[ERROR] Could not write to {FAILURE_MODES_FILE}: {e}")
        raise


def get_meter_list():
    # Load from device_config.json
    config_path = os.path.join(os.path.dirname(os.path.dirname(
        os.path.dirname(__file__))), 'device_config.json')
    try:
        with open(config_path, 'r') as f:
            # Remove comments if any
            lines = [line for line in f if not line.strip().startswith('//')]
            return json.loads(''.join(lines))
    except Exception:
        return []


def dashboard(request):
    from device_config.models import MeterDevice
    meters = MeterDevice.objects.all()
    failure_modes = load_failure_modes()
    available_modes = ['phase_loss', 'overcurrent', 'bad_pf',
                       'overvoltage', 'reverse_power', 'freq_drift', None]
    return render(request, 'meter_readings/dashboard.html', {
        'page_title': 'Meter Readings Dashboard',
        'meters': meters,
        'failure_modes': failure_modes,
        'available_modes': available_modes
    })


def api_set_failure_mode(request):
    if request.method != 'POST':
        return HttpResponseBadRequest('POST only')
    try:
        data = json.loads(request.body.decode('utf-8'))
        meter_name = data['meter_name']
        mode = data['mode']
        modes = load_failure_modes()
        modes[meter_name] = mode
        save_failure_modes(modes)
        return JsonResponse({'status': 'ok', 'meter_name': meter_name, 'mode': mode})
    except Exception as e:
        return JsonResponse({'status': 'error', 'error': str(e)})


def export_excel(request):
    """Export filtered meter readings to Excel"""
    if not EXCEL_AVAILABLE:
        return HttpResponse('openpyxl not installed. Install with: pip install openpyxl', status=500)
    
    # Get same filters as latest_readings
    view_mode = request.GET.get('view_mode', 'latest')
    selected_meters = request.GET.getlist('meters')
    selected_columns = request.GET.getlist('columns')
    hours_back = request.GET.get('hours', '24')
    
    try:
        hours_back = int(hours_back)
    except ValueError:
        hours_back = 24
    
    db_settings = getattr(settings, 'DATABASES', {}).get('default', {})
    conn = None
    cur = None
    readings = []
    columns = []
    
    try:
        conn = psycopg2.connect(
            dbname=db_settings.get('NAME', 'mfmdb'),
            user=db_settings.get('USER', 'mfmuser'),
            password=db_settings.get('PASSWORD', 'devi'),
            host=db_settings.get('HOST', 'localhost'),
            port=db_settings.get('PORT', 5432)
        )
        cur = conn.cursor()
        
        # Build query (same logic as latest_readings)
        if view_mode == 'latest':
            query = "SELECT DISTINCT ON (meter_name) * FROM meter_readings"
            if selected_meters:
                placeholders = ','.join(['%s'] * len(selected_meters))
                query += f" WHERE meter_name IN ({placeholders})"
            query += " ORDER BY meter_name, time DESC;"
            
            if selected_meters:
                cur.execute(query, selected_meters)
            else:
                cur.execute(query)
        else:  # timeseries
            query = "SELECT * FROM meter_readings WHERE time >= NOW() - INTERVAL '%s hours'"
            params = [hours_back]
            
            if selected_meters:
                placeholders = ','.join(['%s'] * len(selected_meters))
                query += f" AND meter_name IN ({placeholders})"
                params.extend(selected_meters)
            
            query += " ORDER BY time DESC LIMIT 1000;"
            cur.execute(query, params)
        
        readings = cur.fetchall()
        all_columns = [desc[0] for desc in cur.description]
        
        # Filter columns if specified
        if selected_columns:
            col_indices = [i for i, col in enumerate(all_columns) if col in selected_columns]
            columns = [all_columns[i] for i in col_indices]
            readings = [[row[i] for i in col_indices] for row in readings]
        else:
            columns = all_columns
    
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Meter Readings ({view_mode})"
    
    # Header styling
    header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    
    # Write data
    for row_idx, row_data in enumerate(readings, 2):
        for col_idx, value in enumerate(row_data, 1):
            # Convert datetime to string for Excel compatibility
            if isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    filename = f"meter_readings_{view_mode}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

